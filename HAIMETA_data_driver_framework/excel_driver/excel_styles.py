'''
    1. 成功， 显示绿色底，且字体加粗
    2. 失败，显示红色底，且字体加粗
'''

from openpyxl.styles import PatternFill, Font

def pass_(cell):
    cell.value = 'pass'
    cell.fill=PatternFill(patternType='solid',fgColor='00FF00')
    cell.font=Font(bold=True)

def fail_(cell):
    cell.value = 'fail'
    cell.fill=PatternFill(patternType='solid',fgColor='FF0000')
    cell.font=Font(bold=True)
