import traceback
import openpyxl

from HAIMETA_data_driver_framework.config.get_logger import get_logger
from HAIMETA_data_driver_framework.excel_driver.excel_styles import pass_, fail_
from HAIMETA_data_driver_framework.web_keys.web_keys import WebKeys
from HAIMETA_data_driver_framework.config.element_config import ElementLocators

log= get_logger()

success = 0
fail = 0
failed_cases = []
generation_times = {}  # 存储每个成功用例的生图时间
retry_success_cases = {}  # 存储重试成功的用例及其成功的尝试次数

def arguments(data):
    temp_data = {}  # Changed from list to dictionary
    if data:
        str_temp = data.split(";")
        for temp in str_temp:
            t = temp.split("=", 1)
            if len(t) == 2:  # Make sure we have both key and value
                # 移除所有多余的空格，包括前后和中间的空格
                key = "".join(t[0].split())
                value = t[1].strip()
                # 检查是否是ElementLocators中的属性
                if hasattr(ElementLocators, value):
                    temp_data[key] = getattr(ElementLocators, value)
                else:
                    temp_data[key] = value
    return temp_data

def run(file_name):
    global success, fail, failed_cases
    current_sheet = None  # 用于跟踪当前正在处理的sheet名称
    try:
        # 读取excel
        excel = openpyxl.load_workbook(file_name)
        # 考虑多个sheet页面需要执行，所以获取全部sheet页面
        sheets = excel.sheetnames
        for name in sheets:
            current_sheet = name  # 更新当前sheet名称
            # 新增过滤条件：跳过包含_template的sheet和临时文件
            if '_template' in name.lower() or name.startswith('~$'):
                log.info(f'跳过工作表: {name}')
                continue
                
            sheet = excel[name]
            log.info(f'正在执行{name}测试用例')
            for values in sheet.values:
                if type(values[0]) is int:
                    test_data = arguments(values[2])
                    if values[1] == 'open_browser':
                        wk = WebKeys(**test_data)
                    elif values[1] == 'assert_text':
                        max_retries = 3
                        retry_count = 1
                        case_id = f"{file_name}:{name}:{values[0]}"
                        
                        while retry_count <= max_retries:
                            status = getattr(wk, values[1])(**test_data, expected=values[4])
                            if status:
                                pass_(sheet.cell(row=values[0]+2, column=6))
                                success += 1
                                if retry_count > 1:  # 如果是重试成功的情况
                                    retry_success_cases[case_id] = retry_count
                                break
                            else:
                                if retry_count == max_retries:
                                    fail_(sheet.cell(row=values[0]+2, column=6))
                                    fail += 1
                                    failed_cases.append(case_id)
                                else:
                                    log.warning(f"用例 {case_id} 第{retry_count}次执行失败，准备重试")
                            retry_count += 1
                        excel.save(file_name)
                    elif values[1] == 'start_creation':
                        generation_time = getattr(wk, values[1])(**test_data)
                        if generation_time is not None:
                            case_id = f"{file_name}:{name}:{values[0]}"
                            generation_times[case_id] = generation_time
                    else:
                        getattr(wk, values[1])(**test_data)
    except Exception as e:
        log.error(traceback.format_exc())
        fail += 1
        if current_sheet:  # 只在有当前sheet名称时添加到失败列表
            failed_cases.append(file_name+':'+current_sheet)
    finally:
        if 'excel' in locals():  # 确保 excel 变量已定义
            excel.close()


# 在测试结束后输出结果
def sum_info():
    # 构建生图时间信息
    generation_time_info = "\n生图时间统计："
    if generation_times:
        for case_id, time in generation_times.items():
            generation_time_info += f"\n{case_id}: {time}秒"
    else:
        generation_time_info += "\n无生图时间记录"
        
    # 构建重试成功信息
    retry_info = "\n重试成功用例统计："
    if retry_success_cases:
        for case_id, retry_count in retry_success_cases.items():
            retry_info += f"\n{case_id}: 第{retry_count}次尝试成功"
    else:
        retry_info += "\n无重试成功记录"

    # 构建完整的测试报告
    report = f'''
成功用例数：{success}条
失败用例数：{fail}条
失败用例具体情况如下：{failed_cases}{generation_time_info}{retry_info}
    '''
    
    log.warning(report)
    return report
    